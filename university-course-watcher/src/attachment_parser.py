from __future__ import annotations

import io
import logging
import os
import re
import threading
import zipfile
from concurrent.futures import ThreadPoolExecutor
from urllib.parse import urlparse

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

from .http_state import HttpStateCache
from .hwp_parser import extract_hwp_text, extract_hwpx_text
from .network_safety import SafeHttpSession, require_public_http_url
from .utils import DATA_DIR

try:
    from docx import Document
except ImportError:  # pragma: no cover
    Document = None

try:
    from openpyxl import load_workbook
except ImportError:  # pragma: no cover
    load_workbook = None

try:
    from pypdf import PdfReader
except ImportError:  # pragma: no cover
    PdfReader = None

try:
    import pytesseract
    from PIL import Image
except ImportError:  # pragma: no cover
    pytesseract = None
    Image = None

LOGGER = logging.getLogger(__name__)
MAX_ARCHIVE_UNCOMPRESSED_BYTES = 64_000_000
MAX_ARCHIVE_ENTRY_BYTES = 16_000_000
MAX_IMAGE_PIXELS = 40_000_000


class AttachmentParser:
    def __init__(
        self,
        timeout: int = 20,
        max_bytes: int = 8_000_000,
        max_workers: int = 3,
        state_cache: HttpStateCache | None = None,
    ):
        self.timeout = timeout
        self.max_bytes = max_bytes
        self.max_workers = max_workers
        self.max_pdf_pages = max(1, int(os.getenv("PDF_MAX_PAGES", "8")))
        self.state_cache = state_cache or HttpStateCache(DATA_DIR / "course_http_state.json")
        self.thread_local = threading.local()

    def extract_texts(self, urls: list[str]) -> dict[str, str]:
        result: dict[str, str] = {}

        if not urls:
            return result

        iWorkerCount = min(self.max_workers, len(urls))

        with ThreadPoolExecutor(max_workers=iWorkerCount) as executor:
            lstResults = list(executor.map(self._extract_safely, urls))

        for sUrl, sText in zip(urls, lstResults):
            result[sUrl] = sText

        self.state_cache.save()
        return result

    def _extract_safely(self, sUrl: str) -> str:
        try:
            return self.extract_text(sUrl)
        except Exception as exc:
            LOGGER.info("Attachment parse skipped: %s %s", sUrl, exc)
            return ""

    def extract_image_texts(self, urls: list[str]) -> dict[str, str]:
        result: dict[str, str] = {}

        if not urls:
            return result

        worker_count = min(self.max_workers, len(urls))

        with ThreadPoolExecutor(max_workers=worker_count) as executor:
            extracted_texts = list(executor.map(self._extract_image_safely, urls))

        for url, text in zip(urls, extracted_texts):
            result[url] = text

        return result

    def _extract_image_safely(self, url: str) -> str:
        try:
            return self.extract_image_text(url)
        except Exception as exc:
            LOGGER.info("Image OCR skipped: %s %s", url, exc)
            return ""

    def extract_image_text(self, url: str) -> str:
        if pytesseract is None or Image is None:
            return ""

        require_public_http_url(url)
        session = self._session()
        response = session.get(url, timeout=(4, self.timeout), stream=True)

        response.raise_for_status()
        content = response.raw.read(self.max_bytes + 1, decode_content=True)

        if len(content) > self.max_bytes:
            LOGGER.info("Image too large, skipped OCR: %s", url)
            return ""

        image = Image.open(io.BytesIO(content))

        if image.width * image.height > MAX_IMAGE_PIXELS:
            LOGGER.info("Image pixel count exceeds OCR safety limit: %s", url)
            return ""

        image.load()

        if image.width < 300 or image.height < 150:
            return ""

        try:
            text = pytesseract.image_to_string(image, lang="kor+eng")
        except Exception as exc:
            if exc.__class__.__name__ == "TesseractNotFoundError":
                LOGGER.info("Tesseract executable unavailable; skipped OCR: %s", url)
                return ""
            if exc.__class__.__name__ != "TesseractError":
                raise

            try:
                text = pytesseract.image_to_string(image, lang="eng")
            except Exception as fallback_exc:
                LOGGER.info("English OCR fallback failed: %s %s", url, fallback_exc)
                return ""

        return text[:12000]

    def _session(self) -> requests.Session:
        session = getattr(self.thread_local, "session", None)

        if session is None:
            session = SafeHttpSession()
            session.headers.update({"User-Agent": "Mozilla/5.0 university-course-watcher/1.0"})
            retry = Retry(
                total=2,
                connect=1,
                read=1,
                status=2,
                backoff_factor=0.5,
                status_forcelist=[429, 500, 502, 503, 504],
                allowed_methods=["GET"],
                raise_on_status=False,
            )
            session.mount("http://", HTTPAdapter(max_retries=retry))
            session.mount("https://", HTTPAdapter(max_retries=retry))
            self.thread_local.session = session

        return session

    def extract_text(self, url: str) -> str:
        require_public_http_url(url)
        sSuffix = self._suffix(url)
        if sSuffix == ".zip":
            return ""

        session = self._session()
        dictHeaders = self.state_cache.conditional_headers(url)
        response = session.get(url, headers=dictHeaders, timeout=(4, self.timeout), stream=True)

        if response.status_code == 304:
            return self.state_cache.cached_value(url, "extracted_text")

        response.raise_for_status()
        sSuffix = self._detect_suffix(url, response)
        if sSuffix == ".zip":
            return ""

        bytesContent = response.raw.read(self.max_bytes + 1, decode_content=True)
        if len(bytesContent) > self.max_bytes:
            LOGGER.info("Attachment too large, skipped text extraction: %s", url)
            self.state_cache.update(url, response.headers, bytesContent[: self.max_bytes], extracted_text="")
            return ""

        if self.state_cache.content_matches(url, bytesContent):
            return self.state_cache.cached_value(url, "extracted_text")

        data = io.BytesIO(bytesContent)
        sExtractedText = ""

        if sSuffix == ".pdf" and PdfReader is not None:
            sExtractedText = self._pdf_text(data)
        elif sSuffix == ".docx" and Document is not None:
            if not self._is_safe_zip_archive(bytesContent):
                return ""
            sExtractedText = self._docx_text(data)
        elif sSuffix == ".xlsx" and load_workbook is not None:
            if not self._is_safe_zip_archive(bytesContent):
                return ""
            sExtractedText = self._xlsx_text(data)
        elif sSuffix == ".hwpx":
            sExtractedText = extract_hwpx_text(bytesContent)
        elif sSuffix == ".hwp":
            sExtractedText = extract_hwp_text(bytesContent)

        self.state_cache.update(url, response.headers, bytesContent, extracted_text=sExtractedText)
        return sExtractedText

    def _is_safe_zip_archive(self, content: bytes) -> bool:
        try:
            with zipfile.ZipFile(io.BytesIO(content)) as archive:
                total_size = 0

                for entry in archive.infolist():
                    if entry.file_size > MAX_ARCHIVE_ENTRY_BYTES:
                        return False

                    total_size += entry.file_size
                    if total_size > MAX_ARCHIVE_UNCOMPRESSED_BYTES:
                        return False
        except (zipfile.BadZipFile, OSError):
            return False

        return True

    def _detect_suffix(self, url: str, response: requests.Response) -> str:
        sUrlSuffix = self._suffix(url)

        if sUrlSuffix:
            return sUrlSuffix

        sDisposition = response.headers.get("Content-Disposition", "")
        match = re.search(r"filename\*?=(?:UTF-8''|\")?([^\";]+)", sDisposition, re.IGNORECASE)

        if match:
            sFileName = match.group(1).strip()
            sFileSuffix = self._suffix(sFileName)
            if sFileSuffix:
                return sFileSuffix

        sContentType = response.headers.get("Content-Type", "").split(";", 1)[0].strip().lower()
        dictContentTypes = {
            "application/pdf": ".pdf",
            "application/vnd.openxmlformats-officedocument.wordprocessingml.document": ".docx",
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet": ".xlsx",
            "application/x-hwp": ".hwp",
            "application/haansofthwp": ".hwp",
        }

        return dictContentTypes.get(sContentType, "")

    def _suffix(self, url: str) -> str:
        path = urlparse(url).path.lower()
        for suffix in [".pdf", ".hwp", ".hwpx", ".doc", ".docx", ".xls", ".xlsx", ".zip"]:
            if path.endswith(suffix):
                return suffix
        return ""

    def _pdf_text(self, data: io.BytesIO) -> str:
        reader = PdfReader(data)
        iPageCount = len(reader.pages)
        iLeadingPageCount = min(iPageCount, max(1, self.max_pdf_pages - 2))
        lstPageIndexes = list(range(0, iLeadingPageCount))

        if iPageCount > iLeadingPageCount:
            iTrailingStart = max(iLeadingPageCount, iPageCount - 2)
            lstPageIndexes.extend(range(iTrailingStart, iPageCount))

        return "\n".join(reader.pages[iPageIndex].extract_text() or "" for iPageIndex in lstPageIndexes)[:12000]

    def _docx_text(self, data: io.BytesIO) -> str:
        doc = Document(data)
        return "\n".join(p.text for p in doc.paragraphs)[:12000]

    def _xlsx_text(self, data: io.BytesIO) -> str:
        wb = load_workbook(data, read_only=True, data_only=True)
        values: list[str] = []
        for ws in wb.worksheets[:5]:
            for row in ws.iter_rows(max_row=200, values_only=True):
                values.extend(str(cell) for cell in row if cell is not None)
        return "\n".join(values)[:12000]
